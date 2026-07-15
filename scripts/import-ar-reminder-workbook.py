"""Import FYE-month worksheets into AR Reminder through the local API.

The script is dry-run by default. Pass --apply to write. Existing UEN/name pairs
are skipped, so an interrupted import can resume without creating duplicates.
"""

from __future__ import annotations

import argparse
import json
import re
import urllib.error
import urllib.request
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SHEETS = {
    "FYE 31.01.2026": ("January", 1),
    "FYE 28.02.2026": ("February", 2),
    "FYE 31.03.2026": ("March", 3),
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def clean_date(value):
    cleaned = clean(value)
    if cleaned and re.fullmatch(r"\d{4}-\d{2}-\d{2}", cleaned):
        return cleaned, None
    return None, cleaned


def month_dates(month_number: int):
    last_day = monthrange(2026, month_number)[1]
    fye = date(2026, month_number, last_day)
    due_month = month_number + 7
    due_year = 2026 + (due_month - 1) // 12
    due_month = (due_month - 1) % 12 + 1
    due_day = min(last_day, monthrange(due_year, due_month)[1])
    return fye.isoformat(), date(due_year, due_month, due_day).isoformat()


def read_rows(path: Path):
    workbook = load_workbook(path, read_only=False, data_only=True)
    batches = {}
    for sheet_name, (month_name, month_number) in SHEETS.items():
        sheet = workbook[sheet_name]
        fye_date, due_date = month_dates(month_number)
        rows = []
        for values in sheet.iter_rows(min_row=4, values_only=True):
            company = clean(values[1])
            if not company:
                continue
            prepared_date, prepared_status = clean_date(values[4])
            date_of_agm, agm_status = clean_date(values[5])
            sent_date, sent_status = clean_date(values[6])
            received_date, received_status = clean_date(values[7])
            filling_date, filing_status = clean_date(values[8])
            source_notes = [
                f"{label}={status}" for label, status in (
                    ("Report Ready", prepared_status),
                    ("AGM", agm_status),
                    ("To Client", sent_status),
                    ("Signed", received_status),
                    ("AR", filing_status),
                ) if status
            ]
            remarks = clean(values[16])
            if source_notes:
                source_note = "Source status: " + "; ".join(source_notes)
                remarks = f"{remarks} | {source_note}" if remarks else source_note
            row = {
                "entity_name": company,
                "uen": clean(values[2]),
                "fye_month": month_name,
                "fye_year": 2026,
                "fye_date": fye_date,
                "due_date": due_date,
                "reminder_note": clean(values[3]),
                "prepared_date": prepared_date,
                "date_of_agm": date_of_agm,
                "sent_date": sent_date,
                "received_date": received_date,
                "filling_date": filling_date,
                "xbrl": clean(values[9]),
                "software_update": clean(values[10]),
                "dpo": clean(values[11]),
                "ond_ron": clean(values[12]),
                "pic": clean(values[13]),
                "acc_pic": clean(values[14]),
                "tax_pic": clean(values[15]),
                "remarks": remarks,
                "ar_status": clean(values[17]),
                "accounts_status": clean(values[18]),
            }
            rows.append({key: value for key, value in row.items() if value is not None})
        batches[month_name] = rows
    return batches


def request_json(url: str, payload=None):
    body = json.dumps(payload).encode() if payload is not None else None
    request = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json"} if body else {},
        method="POST" if body else "GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            return json.loads(response.read())
    except urllib.error.HTTPError as error:
        raise RuntimeError(f"HTTP {error.code}: {error.read().decode()}") from error


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--base-url", default="http://localhost:3100")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    batches = read_rows(args.workbook)
    expected = {"January": 15, "February": 17, "March": 62}
    actual = {month: len(rows) for month, rows in batches.items()}
    if actual != expected:
        raise RuntimeError(f"Workbook counts changed: expected {expected}, got {actual}")

    existing = {}
    existing_keys = {}
    for month in batches:
        result = request_json(f"{args.base_url}/api/ar-reminder?month={month}&year=2026")
        existing[month] = result.get("total", 0)
        existing_keys[month] = {
            (str(company.get("uen") or "").strip().upper(), str(company.get("entity_name") or "").strip().upper())
            for company in result.get("companies", [])
        }

    print(json.dumps({"mode": "apply" if args.apply else "dry-run", "source": actual, "existing": existing}, indent=2))
    if not args.apply:
        return
    inserted = {month: 0 for month in batches}
    skipped = {month: 0 for month in batches}
    for month, rows in batches.items():
        for row in rows:
            uen = str(row.get("uen") or "").strip().upper()
            name = str(row["entity_name"]).strip().upper()
            if any((uen and uen == existing_uen) or name == existing_name for existing_uen, existing_name in existing_keys[month]):
                skipped[month] += 1
                continue
            result = request_json(f"{args.base_url}/api/ar-reminder", row)
            if not result.get("ok"):
                raise RuntimeError(f"Insert failed for {month} / {row['entity_name']}: {result}")
            inserted[month] += 1
    print(json.dumps({"inserted": inserted, "skipped_existing": skipped, "total_inserted": sum(inserted.values())}, indent=2))


if __name__ == "__main__":
    main()
